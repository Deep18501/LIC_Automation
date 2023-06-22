from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl
import time
import pyautogui

# Initialize Chrome driver
driver = webdriver.Chrome()
driver.maximize_window()
driver.get('https://merchant.licindia.in/merchant/')
time.sleep(2)

# Close the ad
adclose = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/img[2]')
adclose.click()

# Click on the login button
login = driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[2]/p[1]/code/a[1]/img')
login.click()

# Enter login credentials
logId = driver.find_element(By.XPATH, '/html/body/table/tbody/tr[3]/td/form/input[1]')
logId.send_keys('A05968176')
password = driver.find_element(By.XPATH, '/html/body/table/tbody/tr[3]/td/form/input[2]')
password.send_keys('71791314')

# Wait for user to fill the captcha
print('Fill the Captcha')
question = driver.find_element(By.XPATH, '/html/body/table/tbody/tr[3]/td/form/span[4]')
print(question.text)
captcha = int(input("\n0 if done\t "))
while captcha != 0:
    captcha = int(input("\nEnter 0 if done\t "))

# Submit the login form
logSubmit = driver.find_element(By.XPATH, '/html/body/table/tbody/tr[3]/td/form/input[5]')
logSubmit.click()
time.sleep(2)

# Enter cashier login details
cashID = driver.find_element(By.XPATH, '//*[@id="cashierLogin{actionForm.cashierID}"]')
cashID.send_keys('LP')
cashpass = driver.find_element(By.XPATH, '//*[@id="cashierLogin{actionForm.password}"]')
cashpass.send_keys('55565241')
cashIN = driver.find_element(By.XPATH, '/html/body/div/div/div[2]/div/div/table/tbody/tr/td[2]/div/table/tbody/tr/td/div/div/div/span[1]/div/table/tbody/tr/td/div/table/tbody/tr/td/div/div/form/table/tbody/tr[3]/td/input')
cashIN.click()
time.sleep(1)

# Click on premium button
priumbutton = driver.find_element(By.XPATH, '/html/body/div/div/div[2]/div/div/table/tbody/tr/td[1]/div/table/tbody/tr[2]/td/div/span/div/table/tbody/tr[1]/td/a')
priumbutton.click()
time.sleep(1)

# Load the workbook
wb = openpyxl.load_workbook("C:\\Users\\deepa\\Desktop\\Book1.xlsx")
sh1 = wb['Sheet1']
MaxR = sh1.max_row
print(MaxR)

# Loop through rows in the Excel sheet
for i in range(2, MaxR+1):
    polNO = sh1.cell(i, 2).value
    amt = sh1.cell(i, 3).value
    print(polNO, amt)

    # Click on the proceed button
    proceed_button = driver.find_element(By.XPATH, '/html/body/div/div/div[2]/div/div/table/tbody/tr/td[2]/div/table/tbody/tr/td/div/div/div/span[1]/div/table/tbody/tr/td/div/table/tbody/tr/td/div/div/form/input')
    proceed_button.click()

    # Select payment option as Cash only
    select = Select(driver.find_element(By.XPATH, '//*[@id="quickpremiumcombo"]'))
    select.select_by_value('2')

    # Enter policy number and amount
    policy_number = driver.find_element(By.XPATH, '//*[@id="polNum"]')
    policy_number.send_keys(polNO)
    amount = driver.find_element(By.XPATH, '//*[@id="amount"]')
    amount.send_keys(amt)

    # Click on Calculate Premium
    calculate_premium = driver.find_element(By.XPATH, '//*[@id="calculateprem"]')
    calculate_premium.click()
    time.sleep(1)

    # Click on the Submit button
    submit_button = driver.find_element(By.XPATH, '/html/body/div/div/div[2]/div/div/table/tbody/tr/td[2]/div/table/tbody/tr/td/div/div/div/span[1]/div/table/tbody/tr/td/div/table/tbody/tr/td/div/div/form/table/tbody/tr[2]/td/input[2]')
    submit_button.click()
    time.sleep(1)

    # Take screenshot
    pyautogui.screenshot('screenshot' + str(i) + '.png')

    # Go back to the premium page
    driver.execute_script("window.history.go(-1)")
    time.sleep(1)

# Close the browser
driver.quit()
